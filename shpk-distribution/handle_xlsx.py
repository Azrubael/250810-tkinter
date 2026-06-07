import re
import os
from datetime import datetime
import pprint


def clean_full_name(full_name):
    """Очищення і отримання повного імені.
    """
    if not full_name:
        return None
    result = ""
    try:
        result = " ".join(full_name.split()).strip()
    except Exception as e:
        # print(full_name)
        print(f"Помилка при очищенні імені: {e}")
    return result


def date_to_str(date):
    """Перетворення дати в рядок
    """
    result = ""
    if isinstance(date, datetime):
        result = date.strftime("%d.%m.%Y")
    else:
        result = f"{date}"
    return result


def is_shooter(division):
    """Перевірка, чи відповідає рядок з даними підрозділу регулярному виразу для стрільців.
    """
    pattern = r"^(1|2|3|4)\/(1|2|3|4)\/3$"
    match = re.match(pattern, division)
    return match is not None


def is_company_manager(division):
    """Перевірка, чи відповідає рядок з даними підрозділу регулярному виразу для управління роти.
    """
    pattern = r"^упр\ (1|2|3|4)\/3 бо$"
    match = re.match(pattern, division)
    return match is not None


def is_vidzab(division):
    """Перевірка чи відноситься військовослужбовець до відділення забезпечення.
    """
    pattern = r"^від\.заб\.\/3 бо$"
    match = re.match(pattern, division)
    return match is not None


def is_vidzv(division):
    """Перевірка чи відноситься військовослужбовець до відділення зв'язку.
    """
    pattern = r"^від\.зв\./3 бо$"
    match = re.match(pattern, division)
    return match is not None


def is_vidto(division):
    """Перевірка чи відноситься військовослужбовець до відділення технічного обслуговування.
    """
    pattern = r"^від\.то\/3 бо$"
    match = re.match(pattern, division)
    return match is not None


def is_mp(division):
    """Перевірка чи відноситься військовослужбовець до медичного пункту.
    """
    pattern = r"^м.п./3 бо$"
    match = re.match(pattern, division)
    return match is not None


def is_manager(division):
    """Перевірка, чи відповідає рядок з даними підрозділу регулярному виразу для управління частиною.
    """
    pattern = r"^упр 3 бо$"
    match = re.match(pattern, division)
    return match is not None


def get_platoon_and_company(division):
    """Визначення номера взводу та роти по типовому запису підрозділу.
    """
    pattern = r"^(1|2|3|4)\/(1|2|3|4)\/3$"
    match = re.match(pattern, division)
    if match:
        platoon = match.group(1)  # Перший номер (взвод)
        company = match.group(2)  # Другий номер (підрозділ)
        return platoon, company
    return None, None


def get_company_for_management(division):
    """Визначення номера роти для управління по типовому запису підрозділу.
    """
    pattern = r"^упр\ (1|2|3|4)\/3.*$"
    match = re.match(pattern, division)
    if match:
        company = match.group(1)
        return company
    return None


def read_shpk_file(shpk_file_path):
    """Отримання даних з Excel-файлу.
    """
    import openpyxl

    try:
        shpk_wb = openpyxl.load_workbook(shpk_file_path)
        shpk_ws = shpk_wb["ШПС"]
    except Exception as e:
        print(f"Помилка відкриття {shpk_file_path}: {e}")
        return

    shpk_data = {}
    for row in shpk_ws.iter_rows(min_row=4, max_row=630, values_only=True):
        full_name = row[8]               # Стовпчик I - повне ім'я
        department = row[10]             # Стовпчик K - підрозділ
        rank_in_fact = row[7]            # Стовпчик F - звання фактично

        if rank_in_fact and full_name and department:
            cleaned_name = clean_full_name(full_name)
            platoon = None
            company = None
            
            if is_shooter(department):
                platoon, company =  get_platoon_and_company(department)
            elif is_company_manager(department):
                company = get_company_for_management(department)
                platoon = f"упр {company}/3 бо"
            elif is_vidzab(department):
                company = "від.заб./3 бо"
            elif is_vidzv(department):
                company = "від.зв./3 бо"
            elif is_vidto(department):
                company = "від.то/3 бо"
            elif is_mp(department):
                company = "м.п./3 бо"
            elif is_manager(department):
                company = "упр 3 бо"
            else:
                print(f"WARNING: {cleaned_name} has the wrong department data!")

            shpk_data[cleaned_name] = {
                "department": department,
                "platoon" : platoon,
                "company": company,
                "rank_in_fact": rank_in_fact.lower(),
                "assignment": row[20],                      # відрядження
                "hospital": row[21],                        # лікарня
                "vacation_current": date_to_str(row[23]),   # початок поточної відпустки
                "study": row[25],                           # навчання
                "szch": date_to_str(row[26]),               # статус СЗЧ
                "vacation1": date_to_str(row[29])           # чи пішов у І частину щорічної відпустки
            }

    return shpk_data


def internal_distr():
    """Визначення внутрішньої розподілу.
    """
    return {"offi": 0, "serg": 0, "sold": 0}


def make_list_of_companies():
    return {
        "упр 3 бо": {},
        "1": {},
        "2": {},
        "3": {},
        "4": {},
        "від.зв./3 бо": {},
        "від.заб./3 бо": {},
        "від.то/3 бо": {},
        "м.п./3 бо": {}
    }


def calculate_shpk_list(shpk_data):
    """Визначення кількості особового складу по категоріям для складання розподілу.
    """
    total_counter = {"offi": 0, "serg": 0, "sold": 0, "total": 0}
    report_counter = {
        "ППД": internal_distr(),
        "Відпустка": internal_distr(),
        "Шпиталь": internal_distr(),
        "СЗЧ": internal_distr(),
        "Відрядження": internal_distr()
    }
    personnel_distribution = {"ППД": {}, "Відпустка": {}, "Шпиталь": {}, "СЗЧ": {}}

    for full_name in shpk_data:   # Пропускаємо заголовки
        rank = shpk_data[full_name]["rank_in_fact"]
        department = shpk_data[full_name]["department"]

        if rank.endswith("олдат"):
            total_counter["sold"] += 1
        elif rank.endswith("ержант"):
            total_counter["serg"] += 1
        else:
           total_counter["offi"] += 1
        total_counter["total"] += 1

        if shpk_data[full_name]["assignment"] == "ППД":
            personnel_distribution["ППД"][full_name] = [rank, department]
            if rank.endswith("олдат"):
                report_counter["ППД"]["sold"] += 1
            elif rank.endswith("ержант"):
                report_counter["ППД"]["serg"] += 1
            else:
                report_counter["ППД"]["offi"] += 1
        elif shpk_data[full_name]["assignment"] == "КСП":
            if rank.endswith("олдат"):
                report_counter["Відрядження"]["sold"] += 1
            elif rank.endswith("ержант"):
                report_counter["Відрядження"]["serg"] += 1
            else:
                report_counter["Відрядження"]["offi"] += 1
        elif shpk_data[full_name]["assignment"]:
            if rank.endswith("олдат"):
                report_counter["Відрядження"]["sold"] += 1
            elif rank.endswith("ержант"):
                report_counter["Відрядження"]["serg"] += 1
            else:
                report_counter["Відрядження"]["offi"] += 1

        if shpk_data[full_name]["vacation_current"] != "None":
            personnel_distribution["Відпустка"][full_name] = [rank, department]
            if rank.endswith("олдат"):
                report_counter["Відпустка"]["sold"] += 1
            elif rank.endswith("ержант"):
                report_counter["Відпустка"]["serg"] += 1
            else:
                report_counter["Відпустка"]["offi"] += 1
            continue

        if shpk_data[full_name]["hospital"]:
            personnel_distribution["Шпиталь"][full_name] = [rank, department]
            if rank.endswith("олдат"):
                report_counter["Шпиталь"]["sold"] += 1
            elif rank.endswith("ержант"):
                report_counter["Шпиталь"]["serg"] += 1
            else:
                report_counter["Шпиталь"]["offi"] += 1
            continue

        if shpk_data[full_name]["study"]:
            if rank.endswith("олдат"):
                report_counter["Відрядження"]["sold"] += 1                
            elif rank.endswith("ержант"):
                report_counter["Відрядження"]["serg"] += 1
            else:
                report_counter["Відрядження"]["offi"] += 1                
            continue

        # if re.search(r"\d", shpk_data[full_name]["szch"]):
        if shpk_data[full_name]["szch"] != "None":
            personnel_distribution["СЗЧ"][full_name] = [rank, department]
            if rank.endswith("олдат"):
                report_counter["СЗЧ"]["sold"] += 1
            elif rank.endswith("ержант"):
                report_counter["СЗЧ"]["serg"] += 1
            else:
                report_counter["СЗЧ"]["offi"] += 1
 
    for division in report_counter:
        report_counter[division]["total"] = report_counter[division]["offi"] \
            + report_counter[division]["serg"] + report_counter[division]["sold"]

    return total_counter, report_counter, personnel_distribution


def save_report_ppd(total, report, personnel_distribution, \
                    report_file="d:/tmp/звіт_ППД.xlsx"):
    import openpyxl
    from openpyxl.styles import Font

    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    now = datetime.now()
    report_ws.title = now.strftime("%y%m%d")
    cd = now.date().strftime("%d.%m.%Y")
    ct = now.time().strftime("%H:%M")

    pointers = ["ППД", "Відпустка", "Шпиталь", "СЗЧ", "Відрядження"]
    report_ws["A1"] = f"Розподіл особового складу 3бо станом на {ct} {cd}"
    report_ws.append(["Підрозділ", "Офіцери", "Сержанти", "Солдати", "Загалом"])

    for division in report:
        if division not in pointers: continue
        report_ws.append([division, report[division]["offi"], \
            report[division]["serg"], report[division]["sold"], \
            report[division]["total"]])

    report_ws.append(["Підсумок", total["offi"], total["serg"], total["sold"], total["total"]])
    report_ws.append([])

    for p in pointers:
        if p not in personnel_distribution: continue
        report_ws.append([])
        report_ws.append(["","","","",p])
        # Apply bold font to the last cell in that row
        # column 5 is where 'p' was placed
        last_row = report_ws.max_row
        last_cell = report_ws.cell(row=last_row, column=5)  
        last_cell.font = Font(bold=True)
        i = 1
        for name in personnel_distribution[p]:
            report_ws.append(["","","","","",i,
                personnel_distribution[p][name][0],name,
                personnel_distribution[p][name][1]])
            i += 1

    # Вирівнювання по ширині
    for col in range(2, report_ws.max_column + 1):
        max_length = 0
        col_letter = report_ws.cell(row=1, column=col).column_letter
        for cell in report_ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        # Add a little padding
        adjusted_width = max_length + 2
        report_ws.column_dimensions[col_letter].width = adjusted_width        

    directory, filename = os.path.split(report_file)
    new_filename = report_ws.title + "-" +filename
    new_filepath = os.path.join(directory, new_filename)

    try:
        report_wb.save(new_filepath)
        print(f"Розрахункові дані успішно збережено до '{new_filepath}'.")
    except Exception as e:
        print(f"Помилка при збереженні {new_filepath}: {e}")
    return


def update_distribution_report(shpk_data, distribution_file, \
                               output_file="d:/tmp/3бо.xlsx"):
    import openpyxl

    def make_counter():
        """Створення пустого словника для розподілу персоналу."""
        return {
        "За списком": internal_distr(),
        "Відпустка": internal_distr(),
        "Шпиталь": internal_distr(),
        "Навчання": internal_distr(),
        "Відрядження": internal_distr(),
        "КСП": internal_distr(),
        "ВОП": internal_distr(),
        "СЗЧ": internal_distr(),
        "ППД": internal_distr()
    }

    try:
        d_wb = openpyxl.load_workbook(distribution_file)
        d_ws = d_wb["3БО"]
        print(f"Файл {distribution_file} успішно відкрито.")
    except Exception as e:
        print(f"Помилка відкриття {distribution_file}: {e}")
        return

    companies = make_list_of_companies()

    for company in companies:
        companies[company] = make_counter()

    for full_name in shpk_data:
        rank = shpk_data[full_name]["rank_in_fact"]
        company = shpk_data[full_name]["company"]

        if rank.endswith("олдат"):
            companies[company]["За списком"]["sold"] += 1
        elif rank.endswith("ержант"):
            companies[company]["За списком"]["serg"] += 1
        else:
            companies[company]["За списком"]["offi"] += 1

        if shpk_data[full_name]["vacation_current"] != "None":
            if rank.endswith("олдат"):
                companies[company]["Відпустка"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["Відпустка"]["serg"] += 1
            else:
                companies[company]["Відпустка"]["offi"] += 1

        elif shpk_data[full_name]["hospital"]:
            if rank.endswith("олдат"):
                companies[company]["Шпиталь"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["Шпиталь"]["serg"] += 1
            else:
                companies[company]["Шпиталь"]["offi"] += 1

        elif shpk_data[full_name]["study"]:
            if rank.endswith("олдат"):
                companies[company]["Навчання"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["Навчання"]["serg"] += 1
            else:
                companies[company]["Навчання"]["offi"] += 1

        elif shpk_data[full_name]["assignment"] == "ППД":
            if rank.endswith("олдат"):
                companies[company]["ППД"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["ППД"]["serg"] += 1
            else:
                companies[company]["ППД"]["offi"] += 1

        elif shpk_data[full_name]["assignment"] == "КСП":
            if rank.endswith("олдат"):
                companies[company]["КСП"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["КСП"]["serg"] += 1
            else:
                companies[company]["КСП"]["offi"] += 1

        elif shpk_data[full_name]["assignment"] == "ВОП":
            if rank.endswith("олдат"):
                companies[company]["ВОП"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["ВОП"]["serg"] += 1
            else:
                companies[company]["ВОП"]["offi"] += 1

        elif shpk_data[full_name]["assignment"]:
            if rank.endswith("олдат"):
                companies[company]["Відрядження"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["Відрядження"]["serg"] += 1
            else:
                companies[company]["Відрядження"]["offi"] += 1

        # if re.search(r"\d", shpk_data[full_name]["szch"]):
        if shpk_data[full_name]["szch"] != "None":
            if rank.endswith("олдат"):
                companies[company]["СЗЧ"]["sold"] += 1
            elif rank.endswith("ержант"):
                companies[company]["СЗЧ"]["serg"] += 1
            else:
                companies[company]["СЗЧ"]["offi"] += 1

    comp = list(companies.keys())
    r, c = len(comp), 27

    distrib_matrix = [[0 for _ in range(c+1)]  for _ in range(r+1) ]

    for k in range(r):
        distrib_matrix[k][0] = companies[comp[k]]["За списком"]["offi"]
        distrib_matrix[k][1] = companies[comp[k]]["За списком"]["serg"]
        distrib_matrix[k][2] = companies[comp[k]]["За списком"]["sold"]
        distrib_matrix[k][3] = companies[comp[k]]["Відпустка"]["offi"]
        distrib_matrix[k][4] = companies[comp[k]]["Відпустка"]["serg"]
        distrib_matrix[k][5] = companies[comp[k]]["Відпустка"]["sold"]
        distrib_matrix[k][6] = companies[comp[k]]["Шпиталь"]["offi"]
        distrib_matrix[k][7] = companies[comp[k]]["Шпиталь"]["serg"]
        distrib_matrix[k][8] = companies[comp[k]]["Шпиталь"]["sold"]
        distrib_matrix[k][9] = companies[comp[k]]["Навчання"]["offi"]
        distrib_matrix[k][10] = companies[comp[k]]["Навчання"]["serg"]
        distrib_matrix[k][11] = companies[comp[k]]["Навчання"]["sold"]
        distrib_matrix[k][12] = companies[comp[k]]["Відрядження"]["offi"]
        distrib_matrix[k][13] = companies[comp[k]]["Відрядження"]["serg"]
        distrib_matrix[k][14] = companies[comp[k]]["Відрядження"]["sold"]
        distrib_matrix[k][15] = companies[comp[k]]["КСП"]["offi"]
        distrib_matrix[k][16] = companies[comp[k]]["КСП"]["serg"]
        distrib_matrix[k][17] = companies[comp[k]]["КСП"]["sold"]
        distrib_matrix[k][18] = companies[comp[k]]["ВОП"]["offi"]
        distrib_matrix[k][19] = companies[comp[k]]["ВОП"]["serg"]
        distrib_matrix[k][20] = companies[comp[k]]["ВОП"]["sold"]
        distrib_matrix[k][21] = companies[comp[k]]["СЗЧ"]["offi"]
        distrib_matrix[k][22] = companies[comp[k]]["СЗЧ"]["serg"]
        distrib_matrix[k][23] = companies[comp[k]]["СЗЧ"]["sold"]
        distrib_matrix[k][24] = companies[comp[k]]["ППД"]["offi"]
        distrib_matrix[k][25] = companies[comp[k]]["ППД"]["serg"]
        distrib_matrix[k][26] = companies[comp[k]]["ППД"]["sold"]

    start_row, start_col = 3, 6

    for r in range(len(distrib_matrix)-1):
        for c in range(len(distrib_matrix[0])-1):
            if distrib_matrix[r][c] == 0:
                d_ws.cell(row=start_row+r, column=start_col+c).value = ""
            else:
                d_ws.cell(row=start_row+r, column=start_col+c).value = distrib_matrix[r][c]

    now = datetime.now().strftime("%y%m%d")
    if output_file:
        directory, filename = os.path.split(output_file)
        new_filename = now + "-" +filename
        new_filepath = os.path.join(directory, new_filename)
    else:
        print(f"Наче щось відпувається не за планом...\nЗначення змінної output_file: {output_file}.")

    try:
        d_wb.save(new_filepath)
        print(f"Оновлений розподіл персоналу успішно збережено до '{new_filepath}'.")
    except Exception as e:
        print(f"Помилка при збереженні {new_filepath}: {e}")

    return companies


def save_report_vacation1(shpk_data, report_file="d:/tmp/звіт_відпустка_І_черги.xlsx"):
    """Створення звіту про першу частину щорічної відпустки по особовому складу
    """
    import openpyxl

    report_counter = make_list_of_companies()
    total_counter = make_list_of_companies()
    for c in report_counter:
        report_counter[c] = internal_distr()
    report_counter["Загалом"] = internal_distr()
    for t in total_counter:
        total_counter[t] = internal_distr()
    total_counter["Загалом"] = internal_distr()

    for full_name in shpk_data:
        vacation1 = shpk_data[full_name]["vacation1"]

        rank = shpk_data[full_name]["rank_in_fact"]
        company = shpk_data[full_name]["company"]

        if rank.endswith("олдат"):
            total_counter[company]["sold"] += 1
            total_counter["Загалом"]["sold"] += 1
        elif rank.endswith("ержант"):
            total_counter[company]["serg"] += 1
            total_counter["Загалом"]["serg"] += 1
        else:
           total_counter[company]["offi"] += 1
           total_counter["Загалом"]["offi"] += 1

        if vacation1 == "None":
            continue
        if rank.endswith("олдат"):
            report_counter[company]["sold"] += 1
            report_counter["Загалом"]["sold"] += 1
        elif rank.endswith("ержант"):
            report_counter[company]["serg"] += 1
            report_counter["Загалом"]["serg"] += 1
        else:
           report_counter[company]["offi"] += 1
           report_counter["Загалом"]["offi"] += 1

    for c in report_counter:
        total_counter[c]["total"] = total_counter[c]["offi"] + total_counter[c]["serg"] + total_counter[c]["sold"]
        report_counter[c]["total"] = report_counter[c]["offi"] + report_counter[c]["serg"] + report_counter[c]["sold"]
        report_counter[c]["percent"] = 0
        if total_counter[c]["total"] != 0:
            report_counter[c]["percent"] = report_counter[c]["total"] / total_counter[c]["total"] * 100


    now = datetime.now()
    cd = now.date().strftime("%d.%m.%Y")
    time_stamp = now.strftime("%y%m%d")
    directory, filename = os.path.split(report_file)
    new_filename = time_stamp + "-" +filename
    new_filepath = os.path.join(directory, new_filename)

    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws["A1"] = f"Кількість особового складу 3бо, що відгуляла першу частину щорічної відпустки станом на {cd}."
    report_ws.append(["Підрозділ", "За списком", "Відгуляли", "Процент"])

    for c in report_counter:
        by_list = f"{total_counter[c]["total"]}  ({total_counter[c]['offi']}-{total_counter[c]['serg']}-{total_counter[c]['sold']})"
        took_vacation = f"{report_counter[c]['total']}  ({report_counter[c]['offi']}-{report_counter[c]['serg']}-{report_counter[c]['sold']})"
        percent = f"{report_counter[c]['percent']:.1f}%"
        report_ws.append([c, by_list, took_vacation, percent])

    # Вирівнювання по ширині
    for col in range(1, report_ws.max_column + 1):
        max_length = 16
        col_letter = report_ws.cell(row=1, column=col).column_letter
        report_ws.column_dimensions[col_letter].width = max_length  

    try:
        report_wb.save(new_filepath)
        print(f"Звіт про відпустки І черги успішно збережено до '{new_filepath}'.")
    except Exception as e:
        print(f"Помилка при збереженні звіту про відпустки І черги до файлу {new_filepath}: {e}")

    return report_counter


if __name__ == "__main__":
    uploaded_data = read_shpk_file("d:/tmp/ШПС-T0320.xlsx")
    total_counter, report_counter, personnel_distribution = calculate_shpk_list(uploaded_data)
    # r = update_distribution_report(uploaded_data, "d:/Документи/III БАТ/склад 3 БО/260520-3 батальйон_12-00.xlsx","d:/tmp/test.xlsx")
    # r = save_report_vacation1(uploaded_data)
    pprint.pprint(report_counter)
